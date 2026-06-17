import json
import random
from datetime import datetime
from pathlib import Path
from urllib.parse import quote

from openpyxl import load_workbook, Workbook
from playwright.sync_api import sync_playwright, TimeoutError


BASE_DIR = Path(__file__).parent
CONTACT_FILE = BASE_DIR / "contacts.xlsx"
WHATSAPP_URL = "https://web.whatsapp.com"

today = datetime.now().strftime("%Y-%m-%d")

REPORT_JSON = BASE_DIR / f"whatsapp_report_{today}.json"
REPORT_EXCEL = BASE_DIR / f"whatsapp_report_{today}.xlsx"

SCREENSHOT_DIR = BASE_DIR / "screenshots"
SCREENSHOT_DIR.mkdir(exist_ok=True)


def random_delay(page):
    page.wait_for_timeout(random.randint(2000, 5000))


def read_contacts():
    workbook = load_workbook(CONTACT_FILE)
    sheet = workbook.active
    contacts = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        name, phone, message = row

        if not name or not phone:
            continue

        if not message:
            message = "Hi {name}, this is an automated test message."

        contacts.append({
            "name": str(name),
            "phone": str(phone),
            "message": str(message).replace("{name}", str(name))
        })

    return contacts


def extract_last_3_messages(page):
    try:
        page.wait_for_timeout(5000)

        message_elements = page.locator("div.copyable-text span.selectable-text")
        count = message_elements.count()

        print("Total messages found:", count)

        messages = []

        for i in range(count):
            text = message_elements.nth(i).inner_text().strip()
            if text:
                messages.append(text)

        if len(messages) == 0:
            print("Trying fallback selector...")
            message_elements = page.locator("span.selectable-text")
            count = message_elements.count()

            print("Total messages found using fallback:", count)

            for i in range(count):
                text = message_elements.nth(i).inner_text().strip()
                if text:
                    messages.append(text)

        return messages[-3:]

    except Exception as e:
        print("Error while extracting messages:", e)
        return []


def send_message(page, contact):
    result = {
        "name": contact["name"],
        "phone": contact["phone"],
        "message": contact["message"],
        "status": "Failed",
        "error": "",
        "screenshot": "",
        "last_3_messages": []
    }

    try:
        print(f"Processing: {contact['name']} - {contact['phone']}")

        phone = contact["phone"].replace("+", "").replace(" ", "")
        encoded_message = quote(contact["message"])

        chat_url = f"https://web.whatsapp.com/send?phone={phone}&text={encoded_message}"

        page.goto(chat_url)
        page.wait_for_timeout(8000)

        try:
            send_button = page.wait_for_selector("span[data-icon='send']", timeout=10000)
            send_button.click()
            print("Message sent using Send button.")
        except Exception:
            print("Send button not found, pressing Enter instead...")
            page.keyboard.press("Enter")

        page.wait_for_timeout(7000)

        screenshot_path = SCREENSHOT_DIR / f"{contact['name']}_{today}.png"
        page.screenshot(path=str(screenshot_path), full_page=True)

        result["status"] = "Sent"
        result["screenshot"] = str(screenshot_path)
        result["last_3_messages"] = extract_last_3_messages(page)

    except TimeoutError:
        result["error"] = "Unable to open chat or send message."

    except Exception as e:
        result["error"] = str(e)

    return result


def save_reports(results):
    with open(REPORT_JSON, "w", encoding="utf-8") as file:
        json.dump(results, file, indent=4, ensure_ascii=False)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "WhatsApp Report"

    sheet.append([
        "Name",
        "Phone",
        "Message",
        "Status",
        "Error",
        "Screenshot",
        "Last 3 Messages"
    ])

    for item in results:
        sheet.append([
            item["name"],
            item["phone"],
            item["message"],
            item["status"],
            item["error"],
            item["screenshot"],
            " | ".join(item["last_3_messages"])
        ])

    workbook.save(REPORT_EXCEL)


def main():
    if not CONTACT_FILE.exists():
        print("contacts.xlsx file not found.")
        print(f"Please create it here: {CONTACT_FILE}")
        return

    contacts = read_contacts()
    results = []

    with sync_playwright() as p:
        browser = p.chromium.launch_persistent_context(
            user_data_dir=str(BASE_DIR / "whatsapp_session"),
            headless=False
        )

        page = browser.new_page()
        page.goto(WHATSAPP_URL)

        print("Scan QR code if required.")
        print("Waiting for WhatsApp Web to load...")

        page.wait_for_selector("div[role='textbox']", timeout=180000)

        print("WhatsApp Web loaded successfully.")

        for contact in contacts:
            result = send_message(page, contact)
            results.append(result)
            random_delay(page)

        save_reports(results)

        print("Automation completed.")
        print(f"JSON Report: {REPORT_JSON}")
        print(f"Excel Report: {REPORT_EXCEL}")

        browser.close()


if __name__ == "__main__":
    main()